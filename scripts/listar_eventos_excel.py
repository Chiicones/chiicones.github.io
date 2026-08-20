import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CAMPOS = ["Title", "Date", "Category", "Tags", "Local", "Data", "Horario", "Entrada"]

def extrair_campo(texto, campo):
    padrao = rf"^\s*{campo}:\s*(.+)$"
    match = re.search(padrao, texto, re.MULTILINE | re.IGNORECASE)
    return match.group(1).strip() if match else ""

def gerar_excel(pasta="."):
    arquivos_md = sorted(f for f in os.listdir(pasta) if f.endswith(".md"))

    if not arquivos_md:
        print("Nenhum arquivo .md encontrado na pasta.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"

    # Estilos
    cor_header = "4A90D9"
    fonte_header = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    fill_header = PatternFill("solid", start_color=cor_header)
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_esq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    fill_par = PatternFill("solid", start_color="EBF3FB")
    fill_impar = PatternFill("solid", start_color="FFFFFF")

    # Cabeçalhos
    for col, campo in enumerate(CAMPOS, start=1):
        cell = ws.cell(row=1, column=col, value=campo)
        cell.font = fonte_header
        cell.fill = fill_header
        cell.alignment = alinhamento_centro
        cell.border = borda

    ws.row_dimensions[1].height = 22

    # Dados
    for linha, arquivo in enumerate(arquivos_md, start=2):
        caminho = os.path.join(pasta, arquivo)
        with open(caminho, encoding="utf-8-sig") as f:
            conteudo = f.read()

        fill_linha = fill_par if linha % 2 == 0 else fill_impar

        for col, campo in enumerate(CAMPOS, start=1):
            valor = extrair_campo(conteudo, campo)
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = alinhamento_esq
            cell.fill = fill_linha
            cell.border = borda

        ws.row_dimensions[linha].height = 18

    # Larguras das colunas
    larguras = {
        "Title": 40, "Date": 14, "Category": 16, "Tags": 45,
        "Local": 25, "Data": 14, "Horario": 16, "Entrada": 30,
    }
    for col, campo in enumerate(CAMPOS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = larguras[campo]

    # Congelar linha do cabeçalho
    ws.freeze_panes = "A2"

    caminho_saida = os.path.join(pasta, "eventos.xlsx")
    wb.save(caminho_saida)
    print(f"{len(arquivos_md)} evento(s) exportado(s) para: {caminho_saida}")

if __name__ == "__main__":
    pasta_script = os.path.dirname(os.path.abspath(__file__))
    gerar_excel(pasta_script)